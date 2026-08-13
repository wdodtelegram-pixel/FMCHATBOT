

from __future__ import annotations

import logging
import math
import threading
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import config

try:
    import pandas as pd
except ImportError:  # pragma: no cover
    pd = None

logger = logging.getLogger(__name__)


def _is_missing(value: Any) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value))


class _StringAccessor:
    def __init__(self, series: "SimpleSeries"):
        self._series = series

    def strip(self) -> "SimpleSeries":
        return SimpleSeries(
            [
                value.strip()
                if isinstance(value, str)
                else str(value).strip()
                if value is not None
                else ""
                for value in self._series.values
            ],
            name=self._series.name,
        )

    def lower(self) -> "SimpleSeries":
        return SimpleSeries(
            [
                value.lower()
                if isinstance(value, str)
                else str(value).lower()
                if value is not None
                else ""
                for value in self._series.values
            ],
            name=self._series.name,
        )


class _SimpleValueCounts(dict):
    def to_dict(self) -> dict[Any, int]:
        return dict(self)


class SimpleSeries:
    def __init__(self, values: list[Any], name: str | None = None):
        self.values = list(values)
        self.name = name

    def astype(self, dtype: Any) -> "SimpleSeries":
        if dtype is str:
            return SimpleSeries(
                ["" if value is None else str(value) for value in self.values],
                name=self.name,
            )
        raise NotImplementedError(f"astype({dtype}) is not supported")

    @property
    def str(self) -> _StringAccessor:
        return _StringAccessor(self)

    def notna(self) -> list[bool]:
        return [not _is_missing(value) for value in self.values]

    def map(self, func: Any) -> "SimpleSeries":
        return SimpleSeries([func(value) for value in self.values], name=self.name)

    def value_counts(self) -> _SimpleValueCounts:
        return _SimpleValueCounts(Counter(self.values))

    def __eq__(self, other: Any) -> list[bool]:
        return [value == other for value in self.values]

    def __ne__(self, other: Any) -> list[bool]:
        return [value != other for value in self.values]

    def __getitem__(self, key: Any) -> Any:
        if isinstance(key, list):
            if all(isinstance(item, bool) for item in key):
                return SimpleSeries(
                    [value for value, keep in zip(self.values, key) if keep],
                    name=self.name,
                )
            return SimpleSeries([self.values[index] for index in key], name=self.name)
        return self.values[key]

    def __iter__(self):
        return iter(self.values)

    def __len__(self) -> int:
        return len(self.values)

    def __repr__(self) -> str:
        return f"SimpleSeries(name={self.name!r}, values={self.values!r})"


class SimpleDataFrame:
    def __init__(self, rows: list[dict[str, Any]], columns: list[str] | None = None):
        self._rows = [dict(row) for row in rows]
        self.columns = list(columns) if columns is not None else list(rows[0].keys()) if rows else []

    def __len__(self) -> int:
        return len(self._rows)

    @property
    def empty(self) -> bool:
        return len(self._rows) == 0

    def __getitem__(self, key: Any) -> Any:
        if isinstance(key, str):
            return SimpleSeries([row.get(key) for row in self._rows], name=key)
        if isinstance(key, list):
            if len(key) == len(self._rows) and all(isinstance(item, bool) for item in key):
                return SimpleDataFrame([row.copy() for row, keep in zip(self._rows, key) if keep], columns=self.columns)
            return SimpleDataFrame([self._rows[index].copy() for index in key], columns=self.columns)
        raise TypeError("Unsupported DataFrame key type")

    def __setitem__(self, key: str, value: Any) -> None:
        if isinstance(value, SimpleSeries):
            if len(value) != len(self._rows):
                raise ValueError("Length of values does not match number of rows")
            if key not in self.columns:
                self.columns.append(key)
            for row, item in zip(self._rows, value.values):
                row[key] = item
        elif isinstance(value, list):
            if len(value) != len(self._rows):
                raise ValueError("Length of values does not match number of rows")
            if key not in self.columns:
                self.columns.append(key)
            for row, item in zip(self._rows, value):
                row[key] = item
        else:
            if key not in self.columns:
                self.columns.append(key)
            for row in self._rows:
                row[key] = value

    def copy(self) -> "SimpleDataFrame":
        return SimpleDataFrame([row.copy() for row in self._rows], columns=self.columns.copy())

    def iterrows(self):
        return enumerate(self._rows)

    def sort_values(self, key: str, ascending: bool = True) -> "SimpleDataFrame":
        sorted_rows = sorted(
            self._rows,
            key=lambda row: (row.get(key) is None, row.get(key)),
            reverse=not ascending,
        )
        return SimpleDataFrame([row.copy() for row in sorted_rows], columns=self.columns.copy())

    def __repr__(self) -> str:
        return f"SimpleDataFrame(rows={len(self._rows)}, columns={self.columns!r})"


# ---------------------------------------------------------------------------
# Module-level cache
# ---------------------------------------------------------------------------
# PURPOSE: re-reading a workbook on every single Telegram message would be slow
# and would hammer the disk. Instead we keep the last DataFrame in memory
# together with the file's modification time. A lock is used because
# python-telegram-bot can process several updates concurrently, and two threads
# rebuilding the cache at once would be wasteful.
_cache_df: Any | None = None
_cache_mtime: float | None = None
_lock = threading.Lock()


# ---------------------------------------------------------------------------
def _download_from_onedrive(destination: Path) -> bool:
    """
    PURPOSE: optional fallback for when the bot runs somewhere without the
    OneDrive desktop client (a VPS, Raspberry Pi, cloud container).

    A OneDrive "anyone with the link" URL can be turned into a direct-download
    URL by appending "&download=1". We fetch the bytes and write them to the
    same local path the rest of the code already expects, so nothing else in
    the bot has to know where the file came from.

    Returns True on success, False on any failure (the caller then falls back
    to whatever local copy already exists).
    """
    if not config.ONEDRIVE_SHARE_LINK:
        return False

    try:
        import requests

        url = config.ONEDRIVE_SHARE_LINK
        if "download=1" not in url:
            url += ("&" if "?" in url else "?") + "download=1"

        response = requests.get(url, timeout=30)
        response.raise_for_status()

        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_bytes(response.content)
        logger.info("Workbook downloaded from OneDrive share link.")
        return True
    except Exception as exc:
        logger.warning("OneDrive download failed (%s). Using local copy.", exc)
        return False


# ---------------------------------------------------------------------------
def _clean(df: Any) -> Any:
    """
    PURPOSE: turn a messy human-maintained spreadsheet into predictable data.

    Real spreadsheets contain trailing spaces, blank filler rows, and status
    values typed as "in progress", "In Progress " and "ONGOING". Fixing that
    here means every downstream function can do a plain equality check.
    """
    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in config.REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        logger.warning("Workbook is missing expected column(s): %s", missing)
        for col in missing:
            df[col] = ""

    df = df[df[config.COL_CATEGORY].notna()].copy()

    for col in df.columns:
        df[col] = [
            value.strip() if isinstance(value, str) else value
            for value in df[col]
        ]

    def _normalize_status(value: Any) -> str:
        if not _is_missing(value):
            cleaned = str(value).strip()
            return config.STATUS_CANONICAL.get(cleaned.lower(), cleaned)
        return "Yet to Start"

    def _normalize_type(value: Any) -> str:
        if not _is_missing(value):
            cleaned = str(value).strip()
            return config.TYPE_CANONICAL.get(cleaned.lower(), cleaned)
        return "-"

    df[config.COL_STATUS] = [_normalize_status(value) for value in df[config.COL_STATUS]]
    df[config.COL_TYPE] = [_normalize_type(value) for value in df[config.COL_TYPE]]

    return df


# ---------------------------------------------------------------------------
def load_jobs(force: bool = False) -> Any:
    """
    PURPOSE: the main entry point used by every command handler.

    Logic:
      * If a OneDrive link is configured, refresh the local copy first.
      * Compare the file's modification time against the cached one.
      * Only re-read the workbook when the file actually changed (or when
        force=True, used by the /refresh command and the background job).

    This is what makes the bot "real-time": the officer saves in Excel Online,
    OneDrive syncs the file, the mtime changes, and the very next Telegram
    command sees the new data.
    """
    global _cache_df, _cache_mtime

    path = config.EXCEL_PATH

    if config.ONEDRIVE_SHARE_LINK:
        _download_from_onedrive(path)

    if not path.exists():
        raise FileNotFoundError(
            f"Maintenance workbook not found at: {path}\n"
            "Check EXCEL_PATH in config.py or your .env file."
        )

    mtime = path.stat().st_mtime

    with _lock:
        # Fast path -- nothing changed since last time, reuse the cache.
        if not force and _cache_df is not None and _cache_mtime == mtime:
            return _cache_df

        logger.info("Reading workbook from disk: %s", path)
        if pd is not None:
            df = pd.read_excel(path, sheet_name=config.SHEET_NAME, engine="openpyxl")
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(path, data_only=True, read_only=True)
            if config.SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet '{config.SHEET_NAME}' not found in workbook. "
                    f"Available sheets: {', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[config.SHEET_NAME]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Workbook does not contain any rows")

            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            data_rows = []
            for row in rows[1:]:
                if row is None:
                    continue
                data_rows.append(
                    {
                        headers[index]: row[index]
                        for index in range(min(len(headers), len(row)))
                    }
                )
            df = SimpleDataFrame(data_rows, columns=headers)

        df = _clean(df)

        _cache_df = df
        _cache_mtime = mtime
        return df


# ---------------------------------------------------------------------------
def get_jobs_for_category(category_key: str) -> Any:
    """
    PURPOSE: the "Search DataFrame" stage of your workflow.

    Given a command key such as "acmv", look up the matching Excel category
    string from config.CATEGORIES and return only those rows. Matching is
    case-insensitive so "acmv", "ACMV" and "Acmv" in the sheet all match.
    """
    df = load_jobs()
    excel_name = config.CATEGORIES[category_key]["excel"]

    mask = df[config.COL_CATEGORY].astype(str).str.strip().str.lower() == excel_name.lower()
    return df[mask]


# ---------------------------------------------------------------------------
def get_summary_counts() -> dict:
    """
    PURPOSE: feed the /service dashboard.

    Returns a dictionary shaped like:
        { "acmv": {"total": 4, "Yet to Start": 1, "In Progress": 2, "Completed": 1}, ... }

    Doing the counting here keeps the message-building code in formatters.py
    purely about presentation.
    """
    df = load_jobs()
    result = {}

    for key, meta in config.CATEGORIES.items():
        rows = df[
            df[config.COL_CATEGORY].astype(str).str.strip().str.lower()
            == meta["excel"].lower()
        ]
        counts = rows[config.COL_STATUS].value_counts().to_dict()
        result[key] = {
            "total": len(rows),
            "Yet to Start": counts.get("Yet to Start", 0),
            "In Progress": counts.get("In Progress", 0),
            "Completed": counts.get("Completed", 0),
        }

    return result
