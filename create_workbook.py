"""
create_workbook.py
==================
PURPOSE
    Generates the starter maintenance_jobs.xlsx that the bot reads.

    Run this ONCE, then upload the resulting file to OneDrive and let the
    desktop client sync it back down. After that, the FM Officer edits it in
    Excel Online and never needs this script again.

    The workbook it produces has:
      * a "Jobs" sheet with the exact column headers config.py expects
      * data validation dropdowns on Category, Maintenance Type and Status, so
        staff cannot introduce a spelling the bot won't recognise
      * a "Guide" sheet explaining the rules to whoever maintains the file

RUN WITH
    python create_workbook.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config

# ---------------------------------------------------------------------------
# Example rows
# ---------------------------------------------------------------------------
# PURPOSE: realistic starter data so you can test every command before any real
# job records exist. Delete these rows once the sheet goes live.
# Column order MUST match HEADERS below.
SAMPLE_ROWS = [
    # Job ID, Category, Description, Location, Type, Status, Est Completion, Next Job
    ("ACMV-001", "ACMV", "AHU filter replacement", "Level 3 Plant Room", "Preventive", "In Progress", "Week 1 January 2026", "Week 1 April 2026"),
    ("ACMV-002", "ACMV", "Chiller quarterly servicing", "Basement 1 Chiller Plant", "Preventive", "Yet to Start", "Week 2 January 2026", "Week 2 April 2026"),
    ("ACMV-003", "ACMV", "FCU water leak rectification", "Level 7 Office", "Corrective", "Completed", "Week 4 December 2025", "-"),
    ("ACMV-004", "ACMV", "Cooling tower chemical dosing", "Roof Level", "Preventive", "In Progress", "Week 3 January 2026", "Week 3 February 2026"),

    ("LFT-001", "Lift & Escalator", "Passenger lift No.1 monthly service", "Tower A Lobby", "Preventive", "Completed", "Week 1 January 2026", "Week 1 February 2026"),
    ("LFT-002", "Lift & Escalator", "Escalator step chain adjustment", "Level 1 to Level 2", "Corrective", "In Progress", "Week 2 January 2026", "-"),
    ("LFT-003", "Lift & Escalator", "Annual load test — service lift", "Tower B Service Core", "Preventive", "Yet to Start", "Week 4 February 2026", "Week 4 February 2027"),

    ("ELE-001", "Electrical", "LV switchboard thermal scanning", "Main Electrical Room", "Preventive", "Yet to Start", "Week 3 January 2026", "Week 3 July 2026"),
    ("ELE-002", "Electrical", "Emergency lighting battery test", "All Staircases", "Preventive", "In Progress", "Week 2 January 2026", "Week 2 April 2026"),
    ("ELE-003", "Electrical", "Faulty ballast replacement", "Level 5 Corridor", "Corrective", "Completed", "Week 1 January 2026", "-"),
    ("ELE-004", "Electrical", "Generator load bank test", "Genset Room", "Preventive", "Yet to Start", "Week 1 March 2026", "Week 1 September 2026"),

    ("FIR-001", "Fire Protection", "Sprinkler head inspection", "Level 1 to Level 10", "Preventive", "In Progress", "Week 2 January 2026", "Week 2 July 2026"),
    ("FIR-002", "Fire Protection", "Fire alarm panel functional test", "Fire Command Centre", "Preventive", "Completed", "Week 1 January 2026", "Week 1 February 2026"),
    ("FIR-003", "Fire Protection", "Hose reel pressure rectification", "Level 4 Riser", "Corrective", "Yet to Start", "Week 3 January 2026", "-"),
    ("FIR-004", "Fire Protection", "Fire pump weekly churn test", "Pump Room B1", "Preventive", "In Progress", "Week 2 January 2026", "Week 3 January 2026"),

    ("PLB-001", "Plumbing & Sanitary", "Water tank cleaning and disinfection", "Roof Tank Room", "Preventive", "Yet to Start", "Week 4 January 2026", "Week 4 July 2026"),
    ("PLB-002", "Plumbing & Sanitary", "Blocked floor trap clearing", "Level 2 Male Toilet", "Corrective", "Completed", "Week 1 January 2026", "-"),
    ("PLB-003", "Plumbing & Sanitary", "Booster pump servicing", "Basement Pump Room", "Preventive", "In Progress", "Week 3 January 2026", "Week 3 April 2026"),
]

HEADERS = [
    config.COL_JOB_ID,
    config.COL_CATEGORY,
    config.COL_DESCRIPTION,
    config.COL_LOCATION,
    config.COL_TYPE,
    config.COL_STATUS,
    config.COL_EST_COMPLETION,
    config.COL_NEXT_JOB,
]

COLUMN_WIDTHS = [12, 22, 38, 28, 20, 16, 24, 24]


def build_workbook(path: Path) -> None:
    """PURPOSE: create the file, style it, and lock down the free-text risk."""
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Jobs -- the sheet the bot actually reads
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = config.SHEET_NAME

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    ws.append(HEADERS)
    for col_index, width in enumerate(COLUMN_WIDTHS, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_index)].width = width

    for row in SAMPLE_ROWS:
        ws.append(list(row))

    # Apply a readable font to the data area.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADERS)):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Freeze the header so it stays visible while scrolling a long job list.
    ws.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Dropdown validation
    # -----------------------------------------------------------------------
    # PURPOSE: this is the single most important quality control in the whole
    # system. If a member of staff types "ongoing" instead of "In Progress",
    # the bot's status counts silently go wrong. Dropdowns make that impossible
    # at the point of entry rather than patching it afterwards.
    last_row = 500  # allow room for future rows

    categories = ",".join(m["excel"] for m in config.CATEGORIES.values())
    dv_category = DataValidation(
        type="list", formula1=f'"{categories}"', allow_blank=False,
        showErrorMessage=True, errorTitle="Invalid category",
        error="Pick a category from the dropdown list.",
    )
    ws.add_data_validation(dv_category)
    dv_category.add(f"B2:B{last_row}")

    dv_type = DataValidation(
        type="list", formula1='"Preventive,Corrective"', allow_blank=False,
        showErrorMessage=True, errorTitle="Invalid type",
        error="Must be Preventive or Corrective.",
    )
    ws.add_data_validation(dv_type)
    dv_type.add(f"E2:E{last_row}")

    dv_status = DataValidation(
        type="list", formula1='"Yet to Start,In Progress,Completed"', allow_blank=False,
        showErrorMessage=True, errorTitle="Invalid status",
        error="Must be Yet to Start, In Progress or Completed.",
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"F2:F{last_row}")

    # -----------------------------------------------------------------------
    # Sheet 2: Guide -- instructions for whoever maintains the workbook
    # -----------------------------------------------------------------------
    # PURPOSE: the person editing this file in six months will not have read
    # any documentation. Putting the rules inside the file itself is the only
    # instruction that reliably survives.
    guide = wb.create_sheet("Guide")
    guide.column_dimensions["A"].width = 100

    guide_lines = [
        ("HOW TO UPDATE THIS WORKBOOK", True),
        ("", False),
        ("This file is read by the Facilities Management Telegram bot.", False),
        ("Edit it in Excel Online. OneDrive syncs your changes automatically and the bot", False),
        ("picks them up within seconds — no restart is needed.", False),
        ("", False),
        ("RULES", True),
        ("1. Only edit the 'Jobs' sheet. Do not rename it or any column header.", False),
        ("2. Category, Maintenance Type and Status all use dropdowns — always use them.", False),
        ("3. Enter dates as text in this exact style:  Week 1 January 2026", False),
        ("   (A real Excel date also works — the bot converts it to the same style.)", False),
        ("4. Put '-' in Next Job Date for one-off corrective jobs that will not repeat.", False),
        ("5. Add new jobs on the next empty row. Never leave a blank row in the middle.", False),
        ("6. To retire a job, delete its row. Do not blank out only some of the cells.", False),
        ("", False),
        ("STATUS MEANINGS", True),
        ("Yet to Start   — scheduled but no work has begun", False),
        ("In Progress    — work has started, not yet handed over", False),
        ("Completed      — work finished and verified", False),
        ("", False),
        ("The sample rows currently in the Jobs sheet are examples. Delete them once", False),
        ("your real job records are entered.", False),
    ]

    for row_index, (text, is_heading) in enumerate(guide_lines, start=1):
        cell = guide.cell(row=row_index, column=1, value=text)
        cell.font = Font(name="Arial", size=11, bold=is_heading)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Workbook created: {path}")
    print(f"Rows written: {len(SAMPLE_ROWS)}")


if __name__ == "__main__":
    build_workbook(config.EXCEL_PATH)
